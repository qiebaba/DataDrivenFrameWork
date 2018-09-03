#封装EXCEL文件的操作方法
from openpyxl import load_workbook
import time

class ParseExcel():
    def __init__(self):
        self.workbook = None
        self.excelFile = None

    #打开excel文件,返回文件对象
    def loadWorkBook(self, excelPathAndName):
        try:
            self.workbook = load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName
        return self.workbook

    #根据sheet名获取sheet对象
    def getSheetByName(self,sheetName):
        try:
            sheet = self.workbook[sheetName]
            print(self.workbook.sheetnames)
            return sheet
        except Exception as e:
            raise e

    #获取结束行号
    def getRowNumber(self,sheet):
        return sheet.max_row

    #获取结束列号
    def getColsNumber(self,sheet):
        return sheet.max_column

    #获取莫一行的数据
    def getRow(self,sheet,rowNo):
        try:
            return sheet[rowNo]
        except Exception as e:
            raise e

    #获取某一列的数据
    def getColumn(self,sheet,colNo):
        try:
            return sheet[colNo]
        except Exception as e:
            raise e

    #根据单元格位置获取单元格的值
    def getCellOfValue(self,sheet,rowNo,colsNo):
        try:
            self.cellValue= sheet.cell(rowNo,colsNo)
            return self.cellValue
        except Exception as e:
            raise e

   #保存excel文件
    def saveExcel(self,excelFile):
        try:
            self.workbook.save(excelFile)
        except Exception as e:
            raise e

    #根据单元格的坐标向单元格写入数据
    def writeCell(self,sheet,content,rowNo = None,colsNo=None):
        try:
            self.getCellOfValue(sheet,rowNo,colsNo)
            self.cellValue.value = content
            self.saveExcel(self.excelFile)
        except Exception as e:
            raise e

    #写入当前时间
    def writeCellCurrentTime(self,sheet,rowNo=None,colsNo=None):
        try:
            now = int(time.time())
            timeArray =time.localtime(now)
            currentTime =time.strftime("%Y-%m-%d%H:%M:%S",timeArray)
            self.cellValue.value = currentTime
            self.saveExcel(self.excelFile)
        except Exception as e:
            raise e

if __name__ == '__main__':
    pe = ParseExcel()
    pe.loadWorkBook(r"C:\Users\Administrator\PycharmProjects\DataDrivenFrameWork\testData\testdata.xlsx")
    sheet = pe.getSheetByName('Sheet1')
    print(pe.getRowNumber(sheet))
    print(pe.getColsNumber(sheet))
    print(pe.getRow(sheet,1))
    print(pe.getColumn(sheet,'A'))
    print(pe.getCellOfValue(sheet,1,2).value)
    pe.writeCell(sheet,"非常好",1,1)
